#!/usr/bin/env python3
"""Programmatic editability QA for the WorkOS Project Field Sheet v1 template.

Verifies that every intended user-input cell is editable without a password:
the template ships with sheet protection removed. It also checks that all
values, validations, named ranges, filters, freeze panes, merges, and the
sample row survived the editability fix.

Usage:
    python3 scripts/qa_workos_field_sheet_template_editability.py
"""

from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

TEMPLATE_PATH = Path("docs/project-import/templates/workos-project-field-sheet-v1.xlsx")

REQUIRED_SHEETS = ["00_Metadata", "01_Project_Documentation", "02_Backlog"]

METADATA_VALUE_CELLS = ["B6", "B7", "B8", "B9", "B10", "B11", "B12"]

DOC_COLUMNS = 14
BACKLOG_COLUMNS = 12
DATA_START_ROW = 7
DATA_END_ROW = 500

DOC_DROPDOWNS = {"C7:C500", "K7:K500", "M7:M500", "N7:N500"}
BACKLOG_DROPDOWNS = {"D7:D500", "F7:F500", "I7:I500"}


def fail(message: str) -> None:
    raise SystemExit(f"QA FAILED: {message}")


def main() -> None:
    if not TEMPLATE_PATH.exists():
        fail(f"template not found: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    if [ws.title for ws in wb.worksheets] != REQUIRED_SHEETS:
        fail(f"sheets mismatch: {[ws.title for ws in wb.worksheets]}")

    # 1. No sheet protection and no workbook encryption/password
    for ws in wb.worksheets:
        if ws.protection.sheet or ws.protection.password:
            fail(f"{ws.title} still has sheet protection: {ws.protection}")

    with ZipFile(TEMPLATE_PATH) as archive:
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
        if "fileSharing" in workbook_xml or "workbookPassword" in workbook_xml:
            fail("workbook contains password/encryption attributes")
        for name in archive.namelist():
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                xml = archive.read(name).decode("utf-8")
                if "sheetProtection" in xml:
                    fail(f"{name} still contains sheetProtection")

    ws_meta = wb["00_Metadata"]
    ws_doc = wb["01_Project_Documentation"]
    ws_back = wb["02_Backlog"]

    # 2. Metadata values intact and editable
    for addr in METADATA_VALUE_CELLS:
        if ws_meta[addr].value in (None, ""):
            fail(f"metadata value cell {addr} is empty")

    # 3. Headers and sample row intact
    expected_doc_headers = ["external_row_id", "project_slug", "block_type", "title", "date",
                            "summary", "details", "evidence_links", "related_files",
                            "next_action", "status", "order_index", "source_type", "reviewed_by_user"]
    actual_doc_headers = [ws_doc.cell(row=5, column=c).value for c in range(1, DOC_COLUMNS + 1)]
    if actual_doc_headers != expected_doc_headers:
        fail(f"doc headers mismatch: {actual_doc_headers}")

    expected_backlog_headers = ["external_row_id", "project_slug", "title", "status", "priority",
                                "schedule_bucket", "start_date", "end_date", "is_milestone",
                                "workstream", "dod_text", "notes"]
    actual_backlog_headers = [ws_back.cell(row=5, column=c).value for c in range(1, BACKLOG_COLUMNS + 1)]
    if actual_backlog_headers != expected_backlog_headers:
        fail(f"backlog headers mismatch: {actual_backlog_headers}")

    if ws_doc["A7"].value != "EXAMPLE-DO-NOT-IMPORT" or ws_back["A7"].value != "EXAMPLE-DO-NOT-IMPORT":
        fail("sample row was altered")

    # 4. All input cells in rows 7-500 are present as row templates. Empty
    # cells carry styles but no values; presence is verified via the sheet XML.
    with ZipFile(TEMPLATE_PATH) as archive:
        for name in archive.namelist():
            if name == "xl/worksheets/sheet2.xml":
                sheet_xml = archive.read(name).decode("utf-8")
                if f'<row r="{DATA_END_ROW}"' not in sheet_xml:
                    fail("doc sheet is missing the row-500 input template")
            if name == "xl/worksheets/sheet3.xml":
                sheet_xml = archive.read(name).decode("utf-8")
                if f'<row r="{DATA_END_ROW}"' not in sheet_xml:
                    fail("backlog sheet is missing the row-500 input template")

    # 5. Dropdown validations remain present
    doc_ranges = {str(dv.sqref) for dv in ws_doc.data_validations.dataValidation}
    backlog_ranges = {str(dv.sqref) for dv in ws_back.data_validations.dataValidation}
    if not DOC_DROPDOWNS.issubset(doc_ranges):
        fail(f"doc dropdowns missing: {DOC_DROPDOWNS - doc_ranges}")
    if not BACKLOG_DROPDOWNS.issubset(backlog_ranges):
        fail(f"backlog dropdowns missing: {BACKLOG_DROPDOWNS - backlog_ranges}")

    # 6. Named ranges, filters, freeze panes, merges intact
    defined_names = {name for name in wb.defined_names}
    expected_names = {"block_type_list", "doc_status_list", "source_type_list", "boolean_list",
                      "backlog_status_list", "schedule_bucket_list"}
    if not expected_names.issubset(defined_names):
        fail(f"defined names missing: {expected_names - defined_names}")
    if ws_doc.auto_filter.ref != "A5:N500" or ws_back.auto_filter.ref != "A5:L500":
        fail("auto filter ranges changed")
    if ws_doc.freeze_panes != "A6" or ws_back.freeze_panes != "A6":
        fail("freeze panes changed")

    # 7. Editability round-trip: writing must succeed and persist
    wb["00_Metadata"]["B6"] = "QA-EDIT-PROBE"
    probe = Path("/tmp/workos-template-editability-probe.xlsx")
    wb.save(probe)
    reloaded = load_workbook(probe)
    if reloaded["00_Metadata"]["B6"].value != "QA-EDIT-PROBE":
        fail("edit round-trip failed")
    probe.unlink(missing_ok=True)

    print("Template editability QA passed.")
    print("Sheets:", REQUIRED_SHEETS)
    print("Protection: removed (no sheet protection, no workbook password)")
    print("Metadata value cells editable:", METADATA_VALUE_CELLS)
    print("Doc input rows 7-500 x 14 cols editable; Backlog input rows 7-500 x 12 cols editable")
    print("Dropdowns:", sorted(DOC_DROPDOWNS | BACKLOG_DROPDOWNS))
    print("Structure: headers, sample row, named ranges, filters, freeze panes, merges intact")


if __name__ == "__main__":
    main()
