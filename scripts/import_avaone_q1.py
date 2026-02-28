import os
import sys
import json
import sqlite3
import re
from datetime import datetime
from openpyxl import load_workbook

# Config
EXCEL_FILE = os.environ.get("EXCEL_FILE", "AVAONE_Strategic_Master_Sheet_Q1_UPDATED.xlsx")
DB_PATH = "data/workos.db"
OUT_DIR = "scripts/out"
OUT_FILE = os.path.join(OUT_DIR, "avaone_q1_payload.json")
MANIFEST_FILE = os.path.join(OUT_DIR, "avaone_q1_manifest.json")

def main():
    mode = os.environ.get("MODE", "create").lower()
    if mode not in ("create", "sync"):
        mode = "create"
        
    top15_only = os.environ.get("TOP15_ONLY") == "1"
    sync_timeline = os.environ.get("SYNC_TIMELINE") == "1"
    
    # 1. Read existing tasks and docs from DB if available
    db_tasks = {}
    control_doc_id = None
    if os.path.exists(DB_PATH):
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("SELECT id, notes FROM tasks WHERE notes LIKE '%project:avaone-q1%'")
            for row in c.fetchall():
                tid, notes = row
                m = re.search(r'(src:xl:\S+:r\d+)', notes or "")
                if m:
                    db_tasks[m.group(1)] = tid
                    
            c.execute("SELECT id FROM docs WHERE content_md LIKE '%project:avaone-q1 control-doc%'")
            row = c.fetchone()
            if row:
                control_doc_id = row[0]
            conn.close()
        except sqlite3.OperationalError:
            pass

    # 2. Load manifest
    manifest = {"src_tags": [], "control_doc_id": None}
    if os.path.exists(MANIFEST_FILE):
        try:
            with open(MANIFEST_FILE, "r") as f:
                manifest = json.load(f)
        except:
            pass
            
    if not control_doc_id and manifest.get("control_doc_id"):
        control_doc_id = manifest.get("control_doc_id")

    # 3. Read Excel
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found. Please place it at the project root.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(EXCEL_FILE, data_only=True)
    
    actions = []
    current_src_tags = []
    
    stats_created = 0
    stats_updated = 0
    stats_removed = 0
    
    sheet_summaries = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Heuristic for table sheet vs reference sheet (at least 2 rows, 2 columns)
        is_table = ws.max_row >= 2 and ws.max_column >= 2
        
        # Find headers
        headers = {}
        header_row = 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            valid_cols = [c for c in row if c is not None and str(c).strip()]
            if len(valid_cols) >= 2:
                header_row = row_idx
                for col_idx, cell in enumerate(row):
                    if cell is not None and str(cell).strip():
                        headers[str(cell).strip().lower()] = col_idx
                break
                
        def get_val(row, col_name_keywords):
            for h, idx in headers.items():
                if any(k.lower() in h for k in col_name_keywords):
                    if idx < len(row):
                        val = row[idx]
                        if val is None: return ""
                        if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
                        return str(val).strip()
            return ""

        if is_table and headers:
            # Row-table sheet
            sheet_created = 0
            sheet_updated = 0
            
            # Special case for top15 only on NanaGarden
            is_nana = "NanaGarden" in sheet_name
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
                # use first column as primary or 'species' if available
                primary_val = get_val(row, ["species"])
                if not primary_val:
                    # fallback to first column
                    if len(row) > 0 and row[0] is not None and str(row[0]).strip():
                        primary_val = str(row[0]).strip()
                    else:
                        continue
                
                tier = get_val(row, ["tier"])
                if is_nana and top15_only:
                    t_lower = tier.lower()
                    if "hero" not in t_lower and "signature" not in t_lower:
                        continue
                
                # Determine title
                title = f"{sheet_name}: {primary_val}"
                if tier:
                    title += f" [Tier:{tier}]"
                    
                src_tag = f"src:xl:{sheet_name.replace(' ', '_')}:r{row_idx}"
                current_src_tags.append(src_tag)
                
                # Extract fields
                size = get_val(row, ["size", "inch"])
                price = get_val(row, ["price", "thb"])
                photos = get_val(row, ["photos ready"])
                title_ready = get_val(row, ["title ready"])
                copy = get_val(row, ["copy ready"])
                url = get_val(row, ["url", "listing"])
                cta = get_val(row, ["cta"])
                status_val = get_val(row, ["status"])
                notes_val = get_val(row, ["notes"])
                
                start_date = get_val(row, ["start date", "scheduled_date", "date"])
                
                notes_text = f"project:avaone-q1\nxl:file:{EXCEL_FILE}\n{src_tag}\n\n"
                
                if size: notes_text += f"Size (inch): {size}\n"
                if price: notes_text += f"Est price range (THB): {price}\n"
                if photos: notes_text += f"Photos ready: {photos}\n"
                if title_ready: notes_text += f"Title ready: {title_ready}\n"
                if copy: notes_text += f"Copy ready: {copy}\n"
                if url: notes_text += f"Listing URL: {url}\n"
                if cta: notes_text += f"CTA: {cta}\n"
                if status_val: notes_text += f"Status: {status_val}\n"
                if notes_val: notes_text += f"Notes: {notes_val}\n"
                
                notes_text += "\n- [ ] Photos ready\n- [ ] Title ready\n- [ ] Copy ready\n- [ ] Listing URL\n- [ ] Status updated"
                
                workspace = "content"
                if "ops" in sheet_name.lower(): workspace = "ops"
                elif "avacrm" in sheet_name.lower(): workspace = "avacrm"
                
                task_status = "inbox"
                task_bucket = "none"
                if start_date:
                    task_status = "planned"
                    task_bucket = "morning"
                    
                task_data = {
                    "title": title,
                    "workspace": workspace,
                    "notes": notes_text
                }
                
                if mode == "create" or src_tag not in db_tasks:
                    task_data["status"] = task_status
                    task_data["schedule_bucket"] = task_bucket
                    task_data["priority"] = 2
                    if sync_timeline and start_date:
                        task_data["scheduled_date"] = start_date[:10]
                    
                    actions.append({
                        "type": "task.create",
                        "data": task_data
                    })
                    sheet_created += 1
                    stats_created += 1
                elif mode == "sync" and src_tag in db_tasks:
                    task_data["id"] = db_tasks[src_tag]
                    if sync_timeline and start_date:
                        task_data["scheduled_date"] = start_date[:10]
                        task_data["status"] = task_status
                        task_data["schedule_bucket"] = task_bucket
                        
                    actions.append({
                        "type": "task.update",
                        "data": task_data
                    })
                    sheet_updated += 1
                    stats_updated += 1
            
            sheet_summaries.append(f"- **{sheet_name}**: {sheet_created} created, {sheet_updated} updated")
            
        else:
            # Reference sheet -> create doc
            doc_title = f"Reference: {sheet_name} - AVAONE Q1"
            doc_content = f"# {sheet_name}\n\nproject:avaone-q1\nxl:file:{EXCEL_FILE}\n\n"
            
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=True):
                doc_content += "| " + " | ".join([str(c).replace('\n', ' ') if c is not None else "" for c in row]) + " |\n"
            
            actions.append({
                "type": "doc.create",
                "data": {
                    "title": doc_title,
                    "content_md": doc_content
                }
            })
            sheet_summaries.append(f"- **{sheet_name}**: Created as Document")

    # Soft-delete processing
    if mode == "sync":
        prev_tags = set(manifest.get("src_tags", []))
        curr_tags = set(current_src_tags)
        removed_tags = prev_tags - curr_tags
        
        for rtag in removed_tags:
            if rtag in db_tasks:
                actions.append({
                    "type": "task.update",
                    "data": {
                        "id": db_tasks[rtag],
                        "status": "done",
                        "notes": f"[SYNC] Removed from Excel on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                    }
                })
                stats_removed += 1

    # Control Doc
    doc_content = f"""# AVAONE Q1 — Control Panel

project:avaone-q1 control-doc

**Source File:** {EXCEL_FILE}
**Import Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
**Mode:** {mode.upper()}

## Stats
- Tasks Created: {stats_created}
- Tasks Updated: {stats_updated}
- Tasks Soft-Deleted: {stats_removed}

## Sheets Processed
""" + "\n".join(sheet_summaries) + """

## Instructions
To re-run sync, place updated Excel file at root and run `./scripts/import_avaone_q1.sh sync`.
"""
    
    if mode == "create" or not control_doc_id:
        actions.insert(0, {
            "type": "doc.create",
            "saveAs": "control_doc",
            "data": {
                "title": "AVAONE Q1 — Control Panel",
                "content_md": doc_content
            }
        })
    else:
        actions.insert(0, {
            "type": "doc.update",
            "data": {
                "id": control_doc_id,
                "title": "AVAONE Q1 — Control Panel",
                "content_md": doc_content
            }
        })

    payload = {"actions": actions}
    
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        
    # Write manifest
    new_manifest = {
        "src_tags": current_src_tags,
        "control_doc_id": control_doc_id # could be None still
    }
    with open(MANIFEST_FILE, "w", encoding="utf-8") as f:
        json.dump(new_manifest, f, indent=2, ensure_ascii=False)
        
    print(json.dumps(payload, indent=2, ensure_ascii=False))
    
    print(f"\n--- Import Stats ({mode.upper()}) ---", file=sys.stderr)
    print(f"Created: {stats_created}", file=sys.stderr)
    print(f"Updated: {stats_updated}", file=sys.stderr)
    print(f"Removed: {stats_removed}", file=sys.stderr)
    print(f"Payload written to {OUT_FILE}", file=sys.stderr)

if __name__ == "__main__":
    main()
